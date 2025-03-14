import re
import pandas as pd
import requests
from lxml import etree
from openpyxl import load_workbook

'''美国拉斯克医学奖获得者'''
headers={
'User-Agent': #'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:134.0) Gecko/20100101 Firefox/134.0'
}

result1=[]
result2=[]
work_dir=r"C:\Users\Lenovo\Desktop\lxy\2award\data"
'''1972'''
# 1946-2024
# for i in range(2024,1945,-1):
for i in range(1972, 1971, -1):
    url='https://laskerfoundation.org/year_of_award/'+str(i)

    response=requests.get(url=url,headers=headers)
    response.encoding='utf-8'
    tree=etree.HTML(response.text)

    acticle_list=tree.xpath('//div[@class="fusion-posts-container fusion-posts-container-infinite"]/article')[1:]
    for acticle in acticle_list:
        '''获奖项目名称'''
        project_name=acticle.xpath('./div[1]/h2/a/text()')[0]
        # print(project_name)
        '''详情页面url'''
        in_url=acticle.xpath('./div[1]/h2/a/@href')[0]
        # print(in_url)
        '''获得时间'''
        time=acticle.xpath('./div[3]/div/a/text()')[0]
        print(time)
        b_response = requests.get(url=in_url, headers=headers)
        b_response.encoding = 'utf-8'
        new_tree = etree.HTML(b_response.text)
        '''奖项名称'''
        award_name=new_tree.xpath('//div[@class="fusion-page-title-wrapper"]/div/h3/text()')[0]
        # print(award_name)
        '''获奖原因'''
        reason_list = new_tree.xpath('//div[@class="fusion-layout-column fusion_builder_column fusion-builder-column-3 fusion_builder_column_2_3 2_3 fusion-flex-column"]/div/div[2]/p//text()')
        if reason_list:
            reason = ''.join(reason_list)
            # print(reason)
        else:
            reason = ''
            # print(reason)

        # divv_list=new_tree.xpath('//div[@class="post-content"]/div[2]/div/div/div/div/div')
        divv_list = new_tree.xpath('//div[@class="post-content"]/div[1]/div/div/div/div/div') #1972
        for div in divv_list:
            '''姓名'''
            name=div.xpath('./div/div[2]/p[@class="aw-name"]/text()')[0]
            print(name)
            '''机构'''
            institution_list = div.xpath('./div/div[2]/p[@class="aw-work"]/text()')
            if institution_list:
                institution = institution_list[0]
                print(institution)
            else:
                institution = ''
                print(institution)

            dict1 = {
                '业务所': '生物经济研究所',
                '标签类别（关注的组织机构及人才类别）': '美国拉斯克医学奖获得者',
                '姓名': name,
                '机构': institution,
                '奖项名称': award_name,
                '获奖类型': '',
                '级别': '',
                '获奖原因': reason,
                '获奖项目名称': project_name,
                '获得时间': time,
                '来源': url
            }
            result1.append(dict1)
            dict1 = {}

            dict2 = {
                '业务所': '生物经济研究所',
                '标签类别（关注的组织机构及人才类别）': '美国拉斯克医学奖获得者',
                '姓名': name,
                '机构': institution,
                '学院': '',
                '性别': '',
                '职称（例如，教授，讲师，长江学者等）': '',
                '电话': '',
                '邮箱': '',
                '出生日期': '',
                '学历': '',
                '学位': '',
                '研究方向': '',
                '工作职务（例：高等学校教师）': '',
                '个人主页': '',
                '个人简介（个人详情）': ''
            }
            result2.append(dict2)
            dict2 = {}
            print('----------------------')

'''存储文件'''
ff=pd.DataFrame(result1)
df=pd.DataFrame(result2)
file_path = f"{work_dir}/data2.xlsx"
ff.to_excel(file_path, sheet_name='奖状字段',index=False,encoding='utf_8')
book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path) as writer:
    writer.book = book
    df.to_excel(writer, sheet_name="采集字段", index=False)

'''续写Excel文件'''
# ff=pd.DataFrame(result1)
# file_path = f"{work_dir}/data1.xlsx"
# d1=pd.read_excel(file_path, sheet_name='奖状字段')
# d1 = pd.concat([d1, ff], ignore_index=True)	# 合并数据
#
# df=pd.DataFrame(result2)
# d2=pd.read_excel(file_path, sheet_name='采集字段')
# d2 = pd.concat([d2, df], ignore_index=True)	# 合并数据
#
# book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
# with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#     writer.book = book
#     d1.to_excel(writer, sheet_name='奖状字段', index=False)  # 重写sheet
#     d2.to_excel(writer, sheet_name='采集字段', index=False)  # 重写sheet

