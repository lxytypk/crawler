import re
import pandas as pd
import requests
from lxml import etree
from openpyxl import load_workbook

'''吴阶平医学奖获得者'''

total_url='https://www.wjpmf.org.cn/award/9/'
origin_url='https://www.wjpmf.org.cn'

headers={
'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
}

result1=[]
result2=[]
work_dir=r"F:\lxy__\pachong\lxy\2award\data"
file_dir=r"F:\lxy__\pachong\lxy\2award"

response=requests.get(url=total_url,headers=headers)
response.encoding='utf-8'
tree=etree.HTML(response.text)

li_list=tree.xpath('//ul[@class="honBox2List list-unstyled list-inline"]/li')
for li in li_list:
    '''获得时间'''
    year=li.xpath('./h3[@class="honYear"]//text()')[0]
    print(year)

    '''姓名'''
    p_list = li.xpath('./div[@class="honItem"]/p')
    for p in p_list:
        '''来源'''
        nn=p.xpath('./a/@href')[0]
        b=origin_url+nn
        print(b)

        name = p.xpath('./a//text()')[0]
        #2024年度吴阶平医学奖获奖者 郭应禄院士
        match = re.search(r'\d{4}(年度|年)吴阶平医学奖(获奖者|获得者)\s(.*?)(院士|教授)', name)
        if match:
            '''姓名'''
            name = match.group(3)
            print(name)
            '''职称（例如，教授，讲师，长江学者等）'''
            job=match.group(4)
            print(job)
        else:
            continue

        b_response = requests.get(url=b, headers=headers)
        b_response.encoding = 'utf-8'
        b_tree = etree.HTML(b_response.text)

        '''个人简介（个人详情）'''
        # profile=''
        # content_list=b_tree.xpath('//div[@class="rightWrap"]/div[2]/p[@style="text-indent: 2em;"]')
        # for content in content_list:
        #     profile+=content.xpath('.//text()')[0]
        # print(profile)


        '''性别'''
        '''出生日期'''
        #1932年11月生
        '''机构'''
        #现任/曾任
        '''工作职务（例：高等学校教师）'''
        '''研究方向'''
        #xxx专家




        dict1 = {
            '业务所': '生物经济研究所',
            '标签类别（关注的组织机构及人才类别）': '吴阶平医学奖获得者',
            '姓名': name,
            '机构': '',
            '奖项名称': '吴阶平医学奖',
            '获奖类型': '',
            '级别': '',
            '获奖原因': '',
            '获奖项目名称': '',
            '获得时间': year,
            '来源': b
        }
        result1.append(dict1)
        dict1 = {}

        dict2 = {
            '业务所': '生物经济研究所',
            '标签类别（关注的组织机构及人才类别）': '吴阶平医学奖获得者',
            '姓名': name,
            '机构': '',
            '学院': '',
            '性别': '',
            '职称（例如，教授，讲师，长江学者等）': job,
            '电话': '',
            '邮箱': '',
            '出生日期': '',
            '学历': '',
            '学位': '',
            '研究方向': '',
            '工作职务（例：高等学校教师）': '',
            '个人主页': '',
            '个人简介（个人详情）': ''
        }
        result2.append(dict2)
        dict2 = {}
        print('-------------------')


'''存储文件'''
ff=pd.DataFrame(result1)
df=pd.DataFrame(result2)
file_path = f"{work_dir}/data9.xlsx"
ff.to_excel(file_path, sheet_name='奖状字段',index=False,encoding='utf_8')
book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path) as writer:
    writer.book = book
    df.to_excel(writer, sheet_name="采集字段", index=False)

'''续写Excel文件'''
ff=pd.DataFrame(result1)
file_path = f"{file_dir}/data.xlsx"
d1=pd.read_excel(file_path, sheet_name='奖状字段')
d1 = pd.concat([d1, ff], ignore_index=True)	# 合并数据

df=pd.DataFrame(result2)
d2=pd.read_excel(file_path, sheet_name='采集字段')
d2 = pd.concat([d2, df], ignore_index=True)	# 合并数据

book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    d1.to_excel(writer, sheet_name='奖状字段', index=False)  # 重写sheet
    d2.to_excel(writer, sheet_name='采集字段', index=False)  # 重写sheet