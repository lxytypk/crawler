import re
import pandas as pd
import requests
from lxml import etree
from openpyxl import load_workbook

'''美国国家科学奖章（生物学）获得者'''

headers={
'User-Agent': #'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:134.0) Gecko/20100101 Firefox/134.0'
}

result1=[]
result2=[]
work_dir=r"C:\Users\Lenovo\Desktop\lxy\2award"

for i in range(0,22):
    url='https://new.nsf.gov/honorary-awards/national-medal-science/recipients?page='+str(i)

    response = requests.get(url=url, headers=headers)
    response.encoding = 'utf-8'
    tree = etree.HTML(response.text)

    div_list=tree.xpath('//div[@class="views-element-container"]/div/div[3]/div')
    for div in div_list:
        '''姓名'''
        first_name = div.xpath('./article/div/div[1]//div[@class="field field-pa-first-name"]/text()')[0]
        last_name = div.xpath('./article/div/div[1]//div[@class="field field-pa-last-name"]/text()')[0]
        name=first_name+' '+last_name
        print(name)
        '''机构'''
        try:
            institution = div.xpath('./article/div/div[1]/div[@class="award-winner-search-result__institutions"]/div/div/text()')[0]
        except IndexError:
            try:
                institution = div.xpath('./article/div/div[1]/div[@class="field field-pa-title"]/text()')[0]
            except IndexError:
                institution = ""
        # print(institution)
        '''职称（例如，教授，讲师，长江学者等）'''
        try:
            profession = div.xpath('./article/div/div[1]/div[@class="field field-pa-title"]/text()')[0]
        except IndexError:
            profession = ""
        # print(profession)
        '''获得时间'''
        year=div.xpath('./article/div/div[2]/div/div/text()')[0]
        print(year)
        '''研究方向'''
        area=div.xpath('./article/div/div[3]/div[@class="award-winner-search-result__research_area"]/div/div/text()')[0]
        # print(area)
        '''获奖原因'''
        reason = div.xpath('./article/div/div[3]/div[@class="award-winner-search-result__citation"]/div/p//text()')[0]
        # print(reason)

        dict1={
            '业务所':'生物经济研究所',
            '标签类别（关注的组织机构及人才类别）':'美国国家科学奖章（生物学）获得者',
            '姓名':name,
            '机构':institution,
            '奖项名称':'National Medal of Science',
            '获奖类型':'',
            '级别':'',
            '获奖原因':reason,
            '获奖项目名称':'',
            '获得时间':year,
            '来源':url
        }
        result1.append(dict1)
        dict1 = {}

        dict2={
            '业务所':'生物经济研究所',
            '标签类别（关注的组织机构及人才类别）':'美国国家科学奖章（生物学）获得者',
            '姓名':name,
            '机构':institution,
            '学院':'',
            '性别':'',
            '职称（例如，教授，讲师，长江学者等）':profession,
            '电话':'',
            '邮箱':'',
            '出生日期':'',
            '学历':'',
            '学位':'',
            '研究方向':area,
            '工作职务（例：高等学校教师）':'',
            '个人主页':'',
            '个人简介（个人详情）':''
        }
        result2.append(dict2)
        dict2 = {}
        print('--------')

'''存储文件'''
# ff=pd.DataFrame(result1)
# df=pd.DataFrame(result2)
# file_path = f"{work_dir}/data4.xlsx"
# ff.to_excel(file_path, sheet_name='奖状字段',index=False,encoding='utf_8')
# book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
# with pd.ExcelWriter(file_path) as writer:
#     writer.book = book
#     df.to_excel(writer, sheet_name="采集字段", index=False)

'''续写Excel文件'''
# ff=pd.DataFrame(result1)
# file_path = f"{work_dir}/data.xlsx"
# d1=pd.read_excel(file_path, sheet_name='奖状字段')
# d1 = pd.concat([d1, ff], ignore_index=True)	# 合并数据
#
# df=pd.DataFrame(result2)
# d2=pd.read_excel(file_path, sheet_name='采集字段')
# d2 = pd.concat([d2, df], ignore_index=True)	# 合并数据
#
# book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
# with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
#     writer.book = book
#     d1.to_excel(writer, sheet_name='奖状字段', index=False)  # 重写sheet
#     d2.to_excel(writer, sheet_name='采集字段', index=False)  # 重写sheet