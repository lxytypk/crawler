import re
import pandas as pd
import requests
from lxml import etree
from openpyxl import load_workbook

'''英国科普利奖章（生物学）获得者'''

headers={
'User-Agent': #'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:134.0) Gecko/20100101 Firefox/134.0'
}

result1=[]
result2=[]
work_dir=r"C:\Users\Lenovo\Desktop\lxy\2award"

'''2024'''
total_url='https://royalsociety.org/medals-and-prizes/copley-medal/'
response = requests.get(url=total_url, headers=headers)
response.encoding = 'utf-8'
tree = etree.HTML(response.text)
'''姓名'''
name=tree.xpath('/html/body/div[1]/main/div[2]/article/section[1]/div/div[2]/div/ul/li/div[2]/h3/text()')[0].strip()
# print(name)
'''获得时间'''
year='2024'
# print(year)
'''获奖原因'''
reason=tree.xpath('/html/body/div[1]/main/div[2]/article/section[1]/div/div[2]/div/ul/li/div[2]/div[@class="expandable-list__desc"]/span/text()')[0]
# print(reason)

dict1={
    '业务所':'生物经济研究所',
    '标签类别（关注的组织机构及人才类别）':'英国科普利奖章（生物学）获得者',
    '姓名':name,
    '机构':'',
    '奖项名称':'The Copley Medal',
    '获奖类型':'',
    '级别':'',
    '获奖原因':reason,
    '获奖项目名称':'',
    '获得时间':year,
    '来源':total_url
}
result1.append(dict1)
dict1 = {}

dict2={
    '业务所':'生物经济研究所',
    '标签类别（关注的组织机构及人才类别）':'英国科普利奖章（生物学）获得者',
    '姓名':name,
    '机构':'',
    '学院':'',
    '性别':'',
    '职称（例如，教授，讲师，长江学者等）':'',
    '电话':'',
    '邮箱':'',
    '出生日期':'',
    '学历':'',
    '学位':'',
    '研究方向':'',
    '工作职务（例：高等学校教师）':'',
    '个人主页':'',
    '个人简介（个人详情）':''
}
result2.append(dict2)
dict2 = {}

def get_data(i):
    url = 'https://royalsociety.org/api/sitecore/PastWinner/PostPastWinner'
    headers = {
        'Referer': 'https://royalsociety.org/medals-and-prizes/copley-medal/',
        'accept': 'application / json, text / plain, * / *',
        'User-Agent':  'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:134.0) Gecko/20100101 Firefox/134.0',
        'Cookie':'shell#lang=en; OptanonAlertBoxClosed=2025-01-18T13:15:37.267Z; OptanonConsent=isGpcEnabled=0&datestamp=Sat+Jan+18+2025+21%3A51%3A28+GMT%2B0800+(%E4%B8%AD%E5%9B%BD%E6%A0%87%E5%87%86%E6%97%B6%E9%97%B4)&version=202401.2.0&browserGpcFlag=0&isIABGlobal=false&hosts=&consentId=84cb2490-98ff-4764-a9da-dfe9cbdf13d0&interactionCount=1&landingPath=NotLandingPage&groups=C0001%3A1%2CC0002%3A0%2CC0004%3A0%2CC0003%3A0&geolocation=JP%3B13&AwaitingReconsent=false'
    }
    value = {
        'dataSourceId': "7cdc5b78-b059-4d41-a5aa-02bf3d8c89d2",
        'page': i, #第几页
    }
    return_text = requests.post(url=url, headers=headers, json=value)
    # print(return_text.text)
    print('-----------------------------------'+str(i)+'-----------------------------------')
    tree = etree.HTML(return_text.text)
    li_list=tree.xpath('//div[@class="section card-list card-list--vertical u-inner"]/div/ul/li')
    for li in li_list:
        '''获得时间'''
        year = li.xpath('./div[@class="expandable-list__text"]/span/strong/text()')[0]
        print(year)
        '''姓名'''
        name = li.xpath('./div[@class="expandable-list__text"]/h3[@class="u-h4 expandable-list__title"]/text()')[0].strip()
        # print(name)
        '''获奖原因'''
        try:
            reason = li.xpath('./div[@class="expandable-list__text"]//span[@class="js-expandableItemText"]/text()')[0]
        except:
            reason = ""
        # reason = li.xpath('./div[@class="expandable-list__text"]//span[@class="js-expandableItemText"]/text()')[0]
        print(reason)

        dict1 = {
            '业务所': '生物经济研究所',
            '标签类别（关注的组织机构及人才类别）': '英国科普利奖章（生物学）获得者',
            '姓名': name,
            '机构': '',
            '奖项名称': 'The Copley Medal',
            '获奖类型': '',
            '级别': '',
            '获奖原因': reason,
            '获奖项目名称': '',
            '获得时间': year,
            '来源': total_url
        }
        result1.append(dict1)
        dict1 = {}

        dict2 = {
            '业务所': '生物经济研究所',
            '标签类别（关注的组织机构及人才类别）': '英国科普利奖章（生物学）获得者',
            '姓名': name,
            '机构': '',
            '学院': '',
            '性别': '',
            '职称（例如，教授，讲师，长江学者等）': '',
            '电话': '',
            '邮箱': '',
            '出生日期': '',
            '学历': '',
            '学位': '',
            '研究方向': '',
            '工作职务（例：高等学校教师）': '',
            '个人主页': '',
            '个人简介（个人详情）': ''
        }
        result2.append(dict2)
        dict2 = {}

for i in range(1,25):
    get_data(i)

'''存储文件'''
# ff=pd.DataFrame(result1)
# df=pd.DataFrame(result2)
# file_path = f"{work_dir}/data5.xlsx"
# ff.to_excel(file_path, sheet_name='奖状字段',index=False,encoding='utf_8')
# book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
# with pd.ExcelWriter(file_path) as writer:
#     writer.book = book
#     df.to_excel(writer, sheet_name="采集字段", index=False)

'''续写Excel文件'''
ff=pd.DataFrame(result1)
file_path = f"{work_dir}/data.xlsx"
d1=pd.read_excel(file_path, sheet_name='奖状字段')
d1 = pd.concat([d1, ff], ignore_index=True)	# 合并数据

df=pd.DataFrame(result2)
d2=pd.read_excel(file_path, sheet_name='采集字段')
d2 = pd.concat([d2, df], ignore_index=True)	# 合并数据

book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    d1.to_excel(writer, sheet_name='奖状字段', index=False)  # 重写sheet
    d2.to_excel(writer, sheet_name='采集字段', index=False)  # 重写sheet