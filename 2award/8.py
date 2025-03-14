import re
import pandas as pd
import requests
from lxml import etree
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By

'''以色列沃尔夫医学奖'''

url='https://wolffund.org.il/the-wolf-prize/#Laureates'

headers={
'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
}

result1=[]
result2=[]
work_dir=r"C:\Users\Lenovo\Desktop\lxy\2award\data"
file_dir=r"C:\Users\Lenovo\Desktop\lxy\2award"

response=requests.get(url=url,headers=headers)
response.encoding='utf-8'
tree=etree.HTML(response.text)

article_list=tree.xpath('//article[@id="post-179"]/div[2]/div[@id="Laureates"]/div[2]/div[@class="posts"]/article')
for article in article_list:
    '''个人简介（个人详情）'''
    b=article.xpath('./a/@href')[0]
    # print(b)
    b_response = requests.get(url=b, headers=headers)
    b_response.encoding = 'utf-8'
    b_tree = etree.HTML(b_response.text)
    '''获得时间'''
    award=b_tree.xpath('//div[@id="excerpt"]/p/text()')[0]
    # print(award)
    match=re.search(r'Wolf Prize Laureate in (.*?) (\d{4})',award)
    type=match.group(1)
    if type!='Medicine':
        continue
    time = match.group(2)
    print(time)
    '''姓名'''
    name=b_tree.xpath('//div[@id="post_details"]/div[@class="small_col mid_col"]/div/h2//text()')[0]
    print(name)

    '''机构'''
    institution_list=b_tree.xpath('//div[@id="post_details"]/div[@class="small_col mid_col"]/div/h4[position() > 1 and following-sibling::p[1]]//text()')
    # print(institution_list)
    institution=''
    found = False
    for r in institution_list:
        if found:
            '''获奖原因'''
            reason=r
            break
        if r.lower() == "award citation:".lower():
            found = True
            continue
        if institution:
            institution += ';'
        institution += r
    print(institution)
    print(reason)
    # print('------------------')

    dict1 = {
        '业务所': '生物经济研究所',
        '标签类别（关注的组织机构及人才类别）': '以色列沃尔夫医学奖',
        '姓名': name,
        '机构': institution,
        '奖项名称': 'Wolf Prize Laureate in Medicine',
        '获奖类型': '',
        '级别': '',
        '获奖原因': reason,
        '获奖项目名称': '',
        '获得时间': time,
        '来源': url
    }
    result1.append(dict1)
    dict1 = {}

    dict2 = {
        '业务所': '生物经济研究所',
        '标签类别（关注的组织机构及人才类别）': '以色列沃尔夫医学奖',
        '姓名': name,
        '机构': institution,
        '学院': '',
        '性别': '',
        '职称（例如，教授，讲师，长江学者等）': '',
        '电话': '',
        '邮箱': '',
        '出生日期': '',
        '学历': '',
        '学位': '',
        '研究方向': '',
        '工作职务（例：高等学校教师）': '',
        '个人主页': '',
        '个人简介（个人详情）': b
    }
    result2.append(dict2)
    dict2 = {}
    if(time=='1978' and name=='George D. Snell'):
        break
    print('--------')

'''存储文件'''
ff=pd.DataFrame(result1)
df=pd.DataFrame(result2)
file_path = f"{work_dir}/data8.xlsx"
ff.to_excel(file_path, sheet_name='奖状字段',index=False,encoding='utf_8')
book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path) as writer:
    writer.book = book
    df.to_excel(writer, sheet_name="采集字段", index=False)

'''续写Excel文件'''
ff=pd.DataFrame(result1)
file_path = f"{file_dir}/data.xlsx"
d1=pd.read_excel(file_path, sheet_name='奖状字段')
d1 = pd.concat([d1, ff], ignore_index=True)	# 合并数据

df=pd.DataFrame(result2)
d2=pd.read_excel(file_path, sheet_name='采集字段')
d2 = pd.concat([d2, df], ignore_index=True)	# 合并数据

book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    d1.to_excel(writer, sheet_name='奖状字段', index=False)  # 重写sheet
    d2.to_excel(writer, sheet_name='采集字段', index=False)  # 重写sheet