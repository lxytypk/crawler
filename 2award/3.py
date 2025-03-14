import re
import time
import pandas as pd
import requests
from lxml import etree
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

'''加拿大盖尔德纳奖获得者'''
headers={
'User-Agent': #'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:134.0) Gecko/20100101 Firefox/134.0'
}

result1=[]
result2=[]
work_dir=r"F:\lxy__\pachong\lxy\2award\data"
file_dir=r"F:\lxy__\pachong\lxy\2award"

# 1959-2024
# for i in range(1995,1958,-1):
# for i in range(2024,1989,-1):
for i in range(1989,1958,-1):
    url='https://www.gairdner.org/winners?date=year-'+str(i)
    print(url)
    options = webdriver.FirefoxOptions()
    options.add_argument('--headless')  # 无头模式
    driver = webdriver.Firefox(options=options)
    driver.get(url)

    # 等待页面加载完成
    WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.ID, "filteredContentResult"))
    )
    # response = requests.get(url=url, headers=headers)
    # response.encoding = 'utf-8'
    # tree = etree.HTML(response.text)
    # print(response.text)
    page_source = driver.page_source
    tree = etree.HTML(page_source)

    li_list=tree.xpath('//*[@id="filteredContentResult"]/li')
    # print(li_list)
    for li in li_list:
        '''姓名'''
        name=li.xpath('./div/div/div[@class="itemContent"]/div/h3/a/text()')[0]
        print(name)
        '''奖项名称'''
        award=li.xpath('./div/div/div[@class="itemContent"]/div/span[@class="awardWon"]/text()')[0]
        print(award)
        # '''职称（例如，教授，讲师，长江学者等）'''
        # try:
        #     profession = li.xpath('./div/div/div[@class="itemContent"]/div/span[@class="designations"]/text()')[0]
        # except IndexError:
        #     profession = ""
        # print(profession)

        b = li.xpath('./div/div/div[@class="itemContent"]/div/h3/a/@href')[0]
        # print(b)
        b_response = requests.get(url=b, headers=headers)
        b_response.encoding = 'utf-8'
        b_tree = etree.HTML(b_response.text)
        '''获奖原因'''
        try:
            reason = b_tree.xpath('//span[@class="quote"]//text()')[0]
        except IndexError:
            reason = ""
        print(reason)
        '''工作职务（例：高等学校教师）'''
        try:
            job = b_tree.xpath('//div[@class="topSection"]/div[@class="left"]/span[@class="position"]//text()')[0]
        except IndexError:
            job = ""
        print(job)
        '''获得时间'''
        year=i
        print(year)
        # '''机构'''
        # pattern = r'(?:University|Institute|Department|Hospital|Centre|Lab|School|College|Chair)\s(?:of|for|at|in)?.*?(?=;|,|$)'
        # # 遍历每条记录提取机构名称
        # for text in job:
        #     institutions = re.findall(pattern, text)
        #     print(institutions)
        dict1 = {
            '业务所': '生物经济研究所',
            '标签类别（关注的组织机构及人才类别）': '加拿大盖尔德纳奖获得者',
            '姓名': name,
            '机构': '',
            '奖项名称': award,
            '获奖类型': '',
            '级别': '',
            '获奖原因': reason,
            '获奖项目名称': '',
            '获得时间': year,
            '来源': url
        }
        result1.append(dict1)
        dict1 = {}

        dict2 = {
            '业务所': '生物经济研究所',
            '标签类别（关注的组织机构及人才类别）': '加拿大盖尔德纳奖获得者',
            '姓名': name,
            '机构': '',
            '学院': '',
            '性别': '',
            '职称（例如，教授，讲师，长江学者等）': '',
            '电话': '',
            '邮箱': '',
            '出生日期': '',
            '学历': '',
            '学位': '',
            '研究方向': '',
            '工作职务（例：高等学校教师）': job,
            '个人主页': '',
            '个人简介（个人详情）': ''
        }
        result2.append(dict2)
        dict2 = {}
        print('----------------------------------------')


driver.quit()

'''存储文件'''
# ff=pd.DataFrame(result1)
# df=pd.DataFrame(result2)
# file_path = f"{work_dir}/data3.xlsx"
# ff.to_excel(file_path, sheet_name='奖状字段',index=False,encoding='utf_8')
# book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
# with pd.ExcelWriter(file_path) as writer:
#     writer.book = book
#     df.to_excel(writer, sheet_name="采集字段", index=False)

'''续写Excel文件'''
ff=pd.DataFrame(result1)
file_path = f"{file_dir}/data.xlsx"
d1=pd.read_excel(file_path, sheet_name='奖状字段')
d1 = pd.concat([d1, ff], ignore_index=True)	# 合并数据

df=pd.DataFrame(result2)
d2=pd.read_excel(file_path, sheet_name='采集字段')
d2 = pd.concat([d2, df], ignore_index=True)	# 合并数据

book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    d1.to_excel(writer, sheet_name='奖状字段', index=False)  # 重写sheet
    d2.to_excel(writer, sheet_name='采集字段', index=False)  # 重写sheet

ff=pd.DataFrame(result1)
file_path = f"{work_dir}/data3.xlsx"
d1=pd.read_excel(file_path, sheet_name='奖状字段')
d1 = pd.concat([d1, ff], ignore_index=True)	# 合并数据

df=pd.DataFrame(result2)
d2=pd.read_excel(file_path, sheet_name='采集字段')
d2 = pd.concat([d2, df], ignore_index=True)	# 合并数据

book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    d1.to_excel(writer, sheet_name='奖状字段', index=False)  # 重写sheet
    d2.to_excel(writer, sheet_name='采集字段', index=False)  # 重写sheet