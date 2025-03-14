import re
import pandas as pd
import requests
from lxml import etree
from openpyxl import load_workbook

'''法国盖伦奖获得者'''

total_url='https://www.galienfoundation.org/laureates-academic-and-public-sector-since-1970'

result1=[]
result2=[]
work_dir=r"F:\lxy__\pachong\lxy\2award\data"
file_dir=r"F:\lxy__\pachong\lxy\2award"

def get_data(i):
    if i==1:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0'
        }
        response = requests.get(url=total_url, headers=headers)
        response.encoding = 'utf-8'
        tree = etree.HTML(response.text)
    # else:
    #     url = 'https://px.ads.linkedin.com/wa/'
    #     headers = {
    #         'Referer': 'https://www.galienfoundation.org/laureates-academic-and-public-sector-since-1970',
    #         'Accept': '*',
    #         'User-Agent':  'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36 Edg/132.0.0.0',
    #         'Cookie':'bcookie="v=2&36a9f179-afcc-4945-89cd-20eca3e5f320"; li_sugr=c0b38bb3-9960-4618-b17b-4cdbe916b4a2; UserMatchHistory=AQJuMJXYvdyg4QAAAZR31oiMFzIc7Ezz652I0HHzTNOosYw9WJDclGPZOZ1AYqzO2G7Ixc0auvNLOA; AnalyticsSyncHistory=AQJk1_PJkBL0NQAAAZR31oiM3V2x6w0sGf6cY3ecLY4H6iPzOfHh0c07nbfraDGNosHbhZc7QuPdAmBpA_OglQ; lidc="b=VGST04:s=V:r=V:a=V:p=V:g=3422:u=1:x=1:i=1737177335:t=1737263735:v=2:sig=AQFt4Ng0oKXnURlxS1o8-Wkj2xfeNqED"; ar_debug=1'
    #     }
    #     value = {
    #         'domAttributes': {
    #             'backgroundImageSrc':'',
    #             'cursor':"pointer",
    #             'elementSemanticType':'',
    #             'elementTitle':'',
    #             'elementType':'',
    #             'elementValue':'',
    #             'imageAlt':'',
    #             'imageSrc':'',
    #             'innerText':i,
    #             'tagName':'A'
    #         }
    #     }
    #
    #     return_text = requests.post(url=url, headers=headers, json=value)
    #     print(return_text.text)
    #
    #     tree = etree.HTML(return_text.text)
    tr_list=tree.xpath('//tbody[@class="rendered_item"]/tr')
    j=1;
    for tr in tr_list:
        print(j)
        j+=1
        '''姓名'''
        try:
            name=tr.xpath('./td[2]/text()')[0]
        except IndexError:
            continue
        print(name)
        '''获得时间'''
        try:
            year=tr.xpath('./td[1]/text()')[0]
        except IndexError:
            year=''
        print(year)
        '''研究方向'''
        try:
            direction=tr.xpath('./td[3]/text()')[0]
        except IndexError:
            direction=''
        print(direction)
        '''获奖项目名称'''
        try:
            project=tr.xpath('./td[4]/text()')[0]
        except IndexError:
            project=''
        print(project)
        '''机构'''
        try:
            organization=tr.xpath('./td[5]/text()')[0]
        except IndexError:
            organization=''
        print(organization)

        dict1 = {
            '业务所': '生物经济研究所',
            '标签类别（关注的组织机构及人才类别）': '法国盖伦奖获得者',
            '姓名': name,
            '机构': organization,
            '奖项名称': '',
            '获奖类型': '',
            '级别': '',
            '获奖原因': '',
            '获奖项目名称': project,
            '获得时间': year,
            '来源': total_url
        }
        result1.append(dict1)
        dict1 = {}

        dict2 = {
            '业务所': '生物经济研究所',
            '标签类别（关注的组织机构及人才类别）': '法国盖伦奖获得者',
            '姓名': name,
            '机构': organization,
            '学院': '',
            '性别': '',
            '职称（例如，教授，讲师，长江学者等）': '',
            '电话': '',
            '邮箱': '',
            '出生日期': '',
            '学历': '',
            '学位': '',
            '研究方向': direction,
            '工作职务（例：高等学校教师）': '',
            '个人主页': '',
            '个人简介（个人详情）': ''
        }
        result2.append(dict2)
        dict2 = {}
        print('------------------------')

    print('-----------------------------------'+str(i)+'-----------------------------------')


get_data(1)

'''存储文件'''
ff=pd.DataFrame(result1)
df=pd.DataFrame(result2)
file_path = f"{work_dir}/data6.xlsx"
ff.to_excel(file_path, sheet_name='奖状字段',index=False,encoding='utf_8')
book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path) as writer:
    writer.book = book
    df.to_excel(writer, sheet_name="采集字段", index=False)

'''续写Excel文件'''
ff=pd.DataFrame(result1)
file_path = f"{file_dir}/data.xlsx"
d1=pd.read_excel(file_path, sheet_name='奖状字段')
d1 = pd.concat([d1, ff], ignore_index=True)	# 合并数据

df=pd.DataFrame(result2)
d2=pd.read_excel(file_path, sheet_name='采集字段')
d2 = pd.concat([d2, df], ignore_index=True)	# 合并数据

book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    d1.to_excel(writer, sheet_name='奖状字段', index=False)  # 重写sheet
    d2.to_excel(writer, sheet_name='采集字段', index=False)  # 重写sheet