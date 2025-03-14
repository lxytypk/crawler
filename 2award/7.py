import re
import time
import pandas as pd
import requests
from lxml import etree
from openpyxl import load_workbook

'''邵逸夫生命科学与医学奖获得者'''
headers={
'User-Agent': #'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0',
    'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:134.0) Gecko/20100101 Firefox/134.0'
}

result1=[]
result2=[]
work_dir=r"F:\lxy__\pachong\lxy\2award\data"
file_dir=r"F:\lxy__\pachong\lxy\2award"

for i in range(1,5):
    if(i==1):
        url='https://www.shawprize.org/sc/prizes-and-laureates/life-science-medicine/'

        response=requests.get(url=url,headers=headers)
        response.encoding='utf-8'
        tree=etree.HTML(response.text)
    else:
        url='https://www.shawprize.org/sc/prizes-and-laureates/life-science-medicine/page/'+str(i)
        response = requests.post(url=url, headers=headers)
        response.encoding = 'utf-8'
        tree = etree.HTML(response.text)
        # print(response.text)

    li_list=tree.xpath('/html/body/div[3]/ul/li')
    # print(li_list)
    for li in li_list:
        '''获得时间'''
        year=li.xpath('./section/div/div[1]//h4//text()')[0]
        print(year)
        '''获奖原因'''
        reason=li.xpath('./section/div/div[3]/div/div//text()')[0]
        print(reason)
        b = li.xpath('./section/div/div[3]/div/a/@href')[0]
        print(b)
        b_response = requests.get(url=b, headers=headers)
        b_response.encoding = 'utf-8'
        b_tree = etree.HTML(b_response.text)
        div_list=b_tree.xpath('/html/body/div[3]/section[1]/div/div[@class="row margin-60-bottom m-margin-8-bottom"]')
        # print(div_list)
        for div in div_list:
            '''姓名'''
            name=div.xpath('./div[@class="col-sm-6 offset-sm-1"]/div/h5//text()')[0]
            print(name)

            content=div.xpath('./div[@class="col-sm-6 offset-sm-1"]/div/div/p/text()')[0].strip()
            print(content)
            '''出生日期'''
            match = re.search(r'(\d{4}年).*?(出生)', content)
            if match:
                birth = match.group(1)
            else:
                birth = ''
            print(birth)
            '''性别'''
            if '他' in content:
                gender='男'
            elif '她' in content:
                gender='女'
            else:
                gender=''
            print(gender)
            '''个人简介（个人详情）'''
            profile=div.xpath('./div[@class="col-sm-6 offset-sm-1"]/div/a/@href')[0]
            print(profile)
            '''工作职务（例：高等学校教师）'''
            match = re.search(r'，(现为|目前是)(.*?)。', content)
            if match:
                job = match.group(2)
            else:
                job = ''
            print(job)
            '''机构'''
            # match = re.search(r'(.*?系|研究院|系主任|讲座教授|主任|讲座|所长|主管|高级|杰出|候任主席|主席|资深研究员).*?', job)
            # if match:
            #     institution = match.group(1)
            # else:
            #     institution = ''
            # print(institution)
            dict1 = {
                '业务所': '生物经济研究所',
                '标签类别（关注的组织机构及人才类别）': '邵逸夫生命科学与医学奖获得者',
                '姓名': name,
                '机构': '',
                '奖项名称': '邵逸夫生命科学与医学奖',
                '获奖类型': '',
                '级别': '',
                '获奖原因': reason,
                '获奖项目名称': '',
                '获得时间': year,
                '来源': url
            }
            result1.append(dict1)
            dict1 = {}

            dict2 = {
                '业务所': '生物经济研究所',
                '标签类别（关注的组织机构及人才类别）': '邵逸夫生命科学与医学奖获得者',
                '姓名': name,
                '机构': '',
                '学院': '',
                '性别': gender,
                '职称（例如，教授，讲师，长江学者等）': '',
                '电话': '',
                '邮箱': '',
                '出生日期': birth,
                '学历': '',
                '学位': '',
                '研究方向': '',
                '工作职务（例：高等学校教师）': job,
                '个人主页': '',
                '个人简介（个人详情）': profile
            }
            result2.append(dict2)
            dict2 = {}
            print('-----------------')

'''存储文件'''
ff=pd.DataFrame(result1)
df=pd.DataFrame(result2)
file_path = f"{work_dir}/data7.xlsx"
ff.to_excel(file_path, sheet_name='奖状字段',index=False,encoding='utf_8')
book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path) as writer:
    writer.book = book
    df.to_excel(writer, sheet_name="采集字段", index=False)

'''续写Excel文件'''
ff=pd.DataFrame(result1)
file_path = f"{file_dir}/data.xlsx"
d1=pd.read_excel(file_path, sheet_name='奖状字段')
d1 = pd.concat([d1, ff], ignore_index=True)	# 合并数据

df=pd.DataFrame(result2)
d2=pd.read_excel(file_path, sheet_name='采集字段')
d2 = pd.concat([d2, df], ignore_index=True)	# 合并数据

book = load_workbook(file_path)  # 该文件必须存在,并且该语句必须在 with pd.ExcelWriter() 之前
with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
    writer.book = book
    d1.to_excel(writer, sheet_name='奖状字段', index=False)  # 重写sheet
    d2.to_excel(writer, sheet_name='采集字段', index=False)  # 重写sheet